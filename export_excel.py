# export_excel.py
from io import BytesIO
from flask import Blueprint, send_file, jsonify
import pandas as pd
import os
import threading
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import uuid

bp = Blueprint('export_excel', __name__)
lock = threading.Lock()

BASE_DIR = os.path.dirname(__file__)
DATA_DIR = os.path.join(BASE_DIR, "data")
EXCEL_PATH = os.path.join(DATA_DIR, "projects.xlsx")
UID_COL = "uid"

DEFAULT_COLUMNS = [
    "No", "BRD No", "Project/Fitur", "Link BRD", "PIC", "Contact Person",
    "Status", "Priority", "Tanggal Submit", "Tanggal Completed", "Catatan"
]

def read_df_safe():
    try:
        os.makedirs(DATA_DIR, exist_ok=True)
        if not os.path.exists(EXCEL_PATH):
            cols = DEFAULT_COLUMNS.copy()
            cols.append(UID_COL)
            return pd.DataFrame(columns=cols)
        df = pd.read_excel(EXCEL_PATH, engine="openpyxl").fillna("")
        for col in DEFAULT_COLUMNS:
            if col not in df.columns:
                df[col] = ""
        if UID_COL not in df.columns:
            df[UID_COL] = [str(uuid.uuid4()) for _ in range(len(df))]
        cols = [c for c in DEFAULT_COLUMNS if c in df.columns]
        if UID_COL not in cols:
            cols.append(UID_COL)
        df = df[cols]
        return df
    except Exception:
        cols = DEFAULT_COLUMNS.copy()
        cols.append(UID_COL)
        return pd.DataFrame(columns=cols)

def _auto_adjust_columns_and_wrap(workbook_bytes):
    workbook_bytes.seek(0)
    wb = load_workbook(workbook_bytes)
    ws = wb.active
    max_width = {}
    for row in ws.iter_rows(values_only=True):
        for idx, cell_value in enumerate(row, start=1):
            text = "" if cell_value is None else str(cell_value)
            length = len(text)
            if any(ord(ch) > 255 for ch in text):
                length = int(length * 1.2)
            max_width[idx] = max(max_width.get(idx, 0), length)
    for idx, width in max_width.items():
        col_letter = get_column_letter(idx)
        adjusted = min(max(8, int(width) + 2), 120)
        try:
            ws.column_dimensions[col_letter].width = adjusted
        except Exception:
            pass
    for row in ws.iter_rows():
        for cell in row:
            try:
                val = cell.value
                if val is not None and isinstance(val, str) and len(val) > 40:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                else:
                    cell.alignment = Alignment(vertical='top')
            except Exception:
                pass
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

@bp.route("/api/projects/export/excel", methods=["GET"])
def export_projects_excel():
    try:
        df = read_df_safe()
        if 'No' in df.columns:
            try:
                df['No'] = range(1, len(df) + 1)
            except Exception:
                pass
        else:
            df.insert(0, 'No', range(1, len(df) + 1))
        for col in df.columns:
            try:
                df[col] = df[col].apply(lambda x: "" if pd.isna(x) else x)
            except Exception:
                pass
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Projects")
        output.seek(0)
        processed = _auto_adjust_columns_and_wrap(output)
        return send_file(
            processed,
            as_attachment=True,
            download_name="projects.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return jsonify({"error": "Export Excel failed", "detail": str(e)}), 500
